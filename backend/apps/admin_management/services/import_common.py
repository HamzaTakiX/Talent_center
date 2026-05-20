"""Shared CSV/Excel parsing utilities for bulk admin imports."""

from __future__ import annotations

import csv
import io
import re
from typing import Any, Optional

MAX_IMPORT_ROWS = 500
ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}


def normalize_header(value: str, aliases: dict[str, str]) -> str:
    key = value.strip().lower().replace(' ', '_')
    key = re.sub(r'[^a-z0-9_àâäéèêëïîôùûüç]', '', key)
    return aliases.get(key, key)


def parse_bool(value: Any) -> bool:
    if value is None or value == '':
        return False
    text = str(value).strip().lower()
    return text in ('1', 'true', 'yes', 'oui', 'y', 'o', 'on')


def parse_int(value: Any) -> Optional[int]:
    if value is None or value == '':
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def parse_csv_list(value: Any) -> list[str]:
    if value is None or value == '':
        return []
    return [part.strip() for part in str(value).split(',') if part.strip()]


def parse_rows_from_csv(*, content: bytes, aliases: dict[str, str]) -> list[dict[str, Any]]:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    headers = [normalize_header(h, aliases) for h in reader.fieldnames]
    rows: list[dict[str, Any]] = []
    for raw in reader:
        row: dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            source_key = reader.fieldnames[idx]
            row[key] = raw.get(source_key, '')
        rows.append(row)
    return rows


def parse_rows_from_xlsx(*, content: bytes, aliases: dict[str, str]) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Excel import requires openpyxl on the server.') from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []

    headers = [normalize_header(str(cell or ''), aliases) for cell in header_row]
    rows: list[dict[str, Any]] = []
    for cells in iterator:
        if not any(cells):
            continue
        row: dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key or idx >= len(cells):
                continue
            row[key] = cells[idx]
        rows.append(row)
    return rows


def parse_import_file(*, content: bytes, filename: str, aliases: dict[str, str]) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith('.csv'):
        return parse_rows_from_csv(content=content, aliases=aliases)
    if lower.endswith('.xlsx'):
        return parse_rows_from_xlsx(content=content, aliases=aliases)
    raise ValueError('Unsupported file type. Use CSV or Excel (.xlsx).')


def validate_import_upload(*, filename: str, rows: list[dict[str, Any]]) -> None:
    lower_name = filename.lower()
    if not any(lower_name.endswith(ext) for ext in ALLOWED_EXTENSIONS):
        raise ValueError('Unsupported file type. Use CSV or Excel (.xlsx).')
    if not rows:
        raise ValueError('The file is empty or has no data rows.')
    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(f'Too many rows (max {MAX_IMPORT_ROWS}).')
