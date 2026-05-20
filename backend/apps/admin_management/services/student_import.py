"""Bulk student import from CSV or Excel files."""

from __future__ import annotations

import csv
import io
import re
from typing import Any, Optional

from django.db import transaction
from django.utils import timezone

from apps.admin_management.models import AcademicYear, ClassGroup, Filiere, ImportLog

from .class_group_import import format_unknown_class_group_error
from .students import create_student

MAX_IMPORT_ROWS = 500
ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}

COLUMN_ALIASES: dict[str, str] = {
    'email': 'email',
    'e-mail': 'email',
    'courriel': 'email',
    'mail': 'email',
    'first_name': 'first_name',
    'firstname': 'first_name',
    'prenom': 'first_name',
    'prénom': 'first_name',
    'last_name': 'last_name',
    'lastname': 'last_name',
    'nom': 'last_name',
    'student_number': 'student_number',
    'studentnumber': 'student_number',
    'numero_etudiant': 'student_number',
    'numéro_etudiant': 'student_number',
    'numero': 'student_number',
    'filiere_id': 'filiere_id',
    'filiere_code': 'filiere_code',
    'filiere': 'filiere_code',
    'programme': 'filiere_code',
    'class_group_id': 'class_group_id',
    'class_group_code': 'class_group_code',
    'class_group': 'class_group_code',
    'classe': 'class_group_code',
    'groupe': 'class_group_code',
    'academic_level_id': 'academic_level_id',
    'academic_year': 'academic_year',
    'academic_year_id': 'academic_year_id',
    'annee_academique': 'academic_year',
    'année_académique': 'academic_year',
    'sso_enabled': 'sso_enabled',
    'sso': 'sso_enabled',
    'grant_access': 'grant_access',
    'acces_plateforme': 'grant_access',
    'platform_access': 'grant_access',
}


def _normalize_header(value: str) -> str:
    key = value.strip().lower().replace(' ', '_')
    key = re.sub(r'[^a-z0-9_àâäéèêëïîôùûüç]', '', key)
    return COLUMN_ALIASES.get(key, key)


def _parse_bool(value: Any) -> bool:
    if value is None or value == '':
        return False
    text = str(value).strip().lower()
    return text in ('1', 'true', 'yes', 'oui', 'y', 'o', 'on')


def _parse_int(value: Any) -> Optional[int]:
    if value is None or value == '':
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _parse_rows_from_csv(content: bytes) -> list[dict[str, Any]]:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    headers = [_normalize_header(h) for h in reader.fieldnames]
    rows: list[dict[str, Any]] = []
    for raw in reader:
        row: dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            source_key = reader.fieldnames[idx]
            row[key] = raw.get(source_key, '')
        rows.append(row)
    return rows


def _parse_rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Excel import requires openpyxl on the server.') from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return []

    headers = [_normalize_header(str(cell or '')) for cell in header_row]
    rows: list[dict[str, Any]] = []
    for cells in iterator:
        if not any(cells):
            continue
        row: dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key or idx >= len(cells):
                continue
            row[key] = cells[idx]
        rows.append(row)
    return rows


def _resolve_academic_ids(row: dict[str, Any]) -> dict[str, Any]:
    filiere_id = _parse_int(row.get('filiere_id'))
    class_group_id = _parse_int(row.get('class_group_id'))
    academic_level_id = _parse_int(row.get('academic_level_id'))
    academic_year_id = _parse_int(row.get('academic_year_id'))
    academic_year = str(row.get('academic_year') or '').strip()

    filiere_code = str(row.get('filiere_code') or '').strip()
    if filiere_code and not filiere_id:
        filiere = Filiere.objects.filter(code__iexact=filiere_code, is_active=True).first()
        if filiere is None:
            raise ValueError(f'Unknown program code: {filiere_code}')
        filiere_id = filiere.id

    class_group_code = str(row.get('class_group_code') or '').strip()
    if class_group_code and not class_group_id:
        class_group = ClassGroup.objects.filter(code__iexact=class_group_code, is_active=True).first()
        if class_group is None:
            class_group = ClassGroup.objects.filter(name__iexact=class_group_code, is_active=True).first()
        if class_group is None:
            raise ValueError(
                format_unknown_class_group_error(
                    class_group_code,
                    filiere_id=filiere_id,
                    filiere_code=filiere_code,
                    academic_year=academic_year,
                ),
            )
        class_group_id = class_group.id

    if academic_year and not academic_year_id:
        year = AcademicYear.objects.filter(code__iexact=academic_year, is_active=True).first()
        if year:
            academic_year_id = year.id

    return {
        'filiere_id': filiere_id,
        'class_group_id': class_group_id,
        'academic_level_id': academic_level_id,
        'academic_year_id': academic_year_id,
        'academic_year': academic_year,
    }


def _row_to_student_kwargs(row: dict[str, Any]) -> dict[str, Any]:
    email = str(row.get('email') or '').strip()
    if not email:
        raise ValueError('Email is required.')

    academic = _resolve_academic_ids(row)
    return {
        'email': email,
        'first_name': str(row.get('first_name') or '').strip(),
        'last_name': str(row.get('last_name') or '').strip(),
        'student_number': str(row.get('student_number') or '').strip(),
        'sso_enabled': _parse_bool(row.get('sso_enabled')),
        'grant_access': _parse_bool(row.get('grant_access')),
        **academic,
    }


def parse_student_import_file(*, content: bytes, filename: str) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith('.csv'):
        return _parse_rows_from_csv(content)
    if lower.endswith('.xlsx'):
        return _parse_rows_from_xlsx(content)
    raise ValueError('Unsupported file type. Use CSV or Excel (.xlsx).')


def import_students_from_file(*, content: bytes, filename: str, created_by) -> dict[str, Any]:
    lower_name = filename.lower()
    if not any(lower_name.endswith(ext) for ext in ALLOWED_EXTENSIONS):
        raise ValueError('Unsupported file type. Use CSV or Excel (.xlsx).')

    rows = parse_student_import_file(content=content, filename=filename)
    if not rows:
        raise ValueError('The file is empty or has no data rows.')
    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(f'Too many rows (max {MAX_IMPORT_ROWS}).')

    log = ImportLog.objects.create(
        import_type=ImportLog.ImportType.STUDENTS,
        status=ImportLog.Status.RUNNING,
        source_filename=filename[:255],
        total_rows=len(rows),
        started_by=created_by,
        started_at=timezone.now(),
    )

    errors: list[dict[str, Any]] = []
    success_count = 0

    for index, row in enumerate(rows, start=2):
        try:
            kwargs = _row_to_student_kwargs(row)
        except ValueError as exc:
            errors.append({'row': index, 'email': str(row.get('email') or ''), 'message': str(exc)})
            continue

        try:
            with transaction.atomic():
                create_student(created_by=created_by, **kwargs)
            success_count += 1
        except ValueError as exc:
            errors.append({'row': index, 'email': kwargs.get('email', ''), 'message': str(exc)})
        except Exception as exc:
            errors.append({'row': index, 'email': kwargs.get('email', ''), 'message': str(exc)})

    error_count = len(errors)
    if success_count == 0 and error_count > 0:
        log_status = ImportLog.Status.FAILED
    elif error_count > 0:
        log_status = ImportLog.Status.PARTIAL
    else:
        log_status = ImportLog.Status.COMPLETED

    log.status = log_status
    log.success_rows = success_count
    log.error_rows = error_count
    log.errors_json = errors
    log.summary_json = {
        'total_rows': len(rows),
        'success_rows': success_count,
        'error_rows': error_count,
    }
    log.completed_at = timezone.now()
    log.save(
        update_fields=[
            'status',
            'success_rows',
            'error_rows',
            'errors_json',
            'summary_json',
            'completed_at',
        ],
    )

    return {
        'import_log_id': log.id,
        'total_rows': len(rows),
        'success_rows': success_count,
        'error_rows': error_count,
        'errors': errors,
        'status': log_status,
    }
