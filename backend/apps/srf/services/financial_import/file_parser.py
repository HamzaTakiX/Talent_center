"""Parse CSV, XLSX, JSON financial import files."""

from __future__ import annotations

import csv
import io
import json
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from typing import Any

from .column_mapping import _normalize_header
from .file_security import MAX_IMPORT_ROWS


def json_safe_value(val: Any) -> Any:
    """Convert parser cell values to JSON/session-serializable primitives."""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    if isinstance(val, timedelta):
        return str(val)
    if isinstance(val, Decimal):
        return str(val)
    if isinstance(val, bytes):
        return val.decode('utf-8', errors='replace')
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float, str)):
        return val
    if isinstance(val, (list, tuple)):
        return [json_safe_value(item) for item in val]
    if isinstance(val, dict):
        return {str(k): json_safe_value(v) for k, v in val.items()}
    return str(val)


def json_safe_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: json_safe_value(value) for key, value in row.items()}


def parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return [], []
    headers = [str(h or '') for h in reader.fieldnames]
    rows: list[dict[str, Any]] = []
    for raw in reader:
        row = {headers[i]: raw.get(reader.fieldnames[i], '') for i in range(len(headers))}
        if any(str(v).strip() for v in row.values()):
            rows.append(row)
    return headers, rows


def parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('Import Excel : openpyxl requis sur le serveur.') from exc

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], []

    headers = [str(cell or '') for cell in header_row]
    rows: list[dict[str, Any]] = []
    for cells in iterator:
        if not any(cells):
            continue
        row: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            val = cells[idx] if idx < len(cells) else ''
            row[header] = json_safe_value(val)
        rows.append(row)
    workbook.close()
    return headers, rows


def parse_json(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    data = json.loads(content.decode('utf-8-sig'))
    if isinstance(data, dict):
        records = data.get('rows') or data.get('data') or data.get('records') or []
    elif isinstance(data, list):
        records = data
    else:
        raise ValueError('JSON invalide : attendu un tableau ou un objet avec "rows".')

    if not records:
        return [], []

    if isinstance(records[0], dict):
        headers = list(records[0].keys())
        return headers, [json_safe_row(record) for record in records]

    raise ValueError('JSON invalide : chaque enregistrement doit être un objet.')


def parse_financial_file(*, content: bytes, filename: str) -> dict[str, Any]:
    lower = filename.lower()
    if lower.endswith('.csv'):
        headers, rows = parse_csv(content)
    elif lower.endswith('.xlsx'):
        headers, rows = parse_xlsx(content)
    elif lower.endswith('.json'):
        headers, rows = parse_json(content)
    else:
        raise ValueError('Type de fichier non supporté.')

    if not rows:
        raise ValueError('Aucune ligne de données dans le fichier.')

    if len(rows) > MAX_IMPORT_ROWS:
        raise ValueError(f'Trop de lignes (maximum {MAX_IMPORT_ROWS}).')

    normalized_headers = [_normalize_header(h) for h in headers]
    safe_rows = [json_safe_row(row) for row in rows]
    return {
        'headers': headers,
        'normalized_headers': normalized_headers,
        'rows': safe_rows,
        'row_count': len(safe_rows),
    }
